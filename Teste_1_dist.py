import pandas as pd
import os
import shutil
import traceback
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Função para converter numpy.datetime64 para datetime
def convert_np_datetime_to_datetime(np_datetime):
    return np_datetime.astype('M8[ms]').astype('O') if not pd.isnull(np_datetime) else None

# Função para formatar a data no formato dd/mm/yyyy
def format_date(cell):
    if isinstance(cell.value, datetime):
        cell.number_format = 'DD/MM/YYYY'
    elif isinstance(cell.value, str):
        try:
            cell.value = datetime.strptime(cell.value, '%Y-%m-%d %H:%M:%S')
            cell.number_format = 'DD/MM/YYYY'
        except ValueError:
            pass
    return cell

# Função para formatar os valores monetários
def format_currency(value):
    try:
        return float(value.replace('.', '').replace(',', '.').strip())
    except ValueError:
        print(f"Erro ao converter valor monetário: {value}")
        return None

# Função para gerar nomes de abas únicos
def generate_unique_sheet_name(base_name, existing_names):
    i = 1
    new_name = f"{base_name}_{i}"
    while new_name in existing_names:
        i += 1
        new_name = f"{base_name}_{i}"
    return new_name

# Função para filtrar ativos pelo nome do cedente e criar/atualizar planilha
def filtrar_cedente_e_atualizar_planilha(nome_cedente, valor_pago, data_pagamento, valor_justo, df_base, df_padrao_path, caminho_base_dados):
    # Verificar os nomes das colunas
    print("Colunas disponíveis no arquivo Excel:", df_base.columns)

    # Filtrar a planilha pelo nome do cedente na coluna "Cedente"
    filtro_cedente = df_base['Cedente'] == nome_cedente
    df_filtrado_cedente = df_base[filtro_cedente]

    # Verificar se encontrou o cedente
    if df_filtrado_cedente.empty:
        print(f"Cedente '{nome_cedente}' não encontrado.")
        return

    # Identificar os códigos de operação em "Ticker 1" e "Ticker 2"
    codigos_operacao = pd.concat([df_filtrado_cedente['Ticker 1'], df_filtrado_cedente['Ticker 2']]).dropna().unique()
    print("Códigos de operação identificados:", codigos_operacao)

    # Filtrar o DataFrame original por esses códigos de operação
    df_filtrado_operacoes = df_base.loc[df_base['Ticker 1'].isin(codigos_operacao) | df_base['Ticker 2'].isin(codigos_operacao)]
    print("Linhas onde os valores em Ticker 1 e Ticker 2 estão presentes na lista de códigos de operação:")
    print(df_filtrado_operacoes)

    # Abrir a aba "Histórico de Captações" e pesquisar os códigos de operação
    df_historico = pd.read_excel(caminho_base_dados, sheet_name='Histórico de Captações')
    valores_historico = []
    for ticker in codigos_operacao:
        valor_ticker = df_historico.loc[df_historico['Token'] == ticker, 'Data de Abertura'].values
        valor_ticker = convert_np_datetime_to_datetime(valor_ticker[0]) if len(valor_ticker) > 0 else None
        valores_historico.append(valor_ticker)

    # Ordenar as datas históricas para garantir que a menor data seja a primeira
    valores_historico.sort()

    # Diretório para salvar a nova planilha ou buscar planilhas existentes
    diretorio_salvar = r"C:\Users\Gabriel Zambello\Hurst Capital\Legal Claims - Documentos\Gestão\Ativos\Controle\Distribuição\Amortização Operações"
    if not os.path.exists(diretorio_salvar):
        os.makedirs(diretorio_salvar)

    # Verificar se existe algum arquivo com os valores de Ticker 1 e Ticker 2 no nome
    arquivos_existentes = [f for f in os.listdir(diretorio_salvar) if any(str(ticker) in f for ticker in codigos_operacao)]

    arquivos_por_ticker = {ticker: None for ticker in codigos_operacao}
    for arquivo in arquivos_existentes:
        for ticker in codigos_operacao:
            if str(ticker) in arquivo:
                arquivos_por_ticker[ticker] = arquivo

    for ticker, arquivo in arquivos_por_ticker.items():
        if arquivo is None:
            # Nenhum arquivo encontrado para este ticker, usar a planilha padrão
            novo_arquivo = os.path.join(diretorio_salvar, f'{ticker}_amortizacao.xlsx')
            shutil.copy(df_padrao_path, novo_arquivo)
            arquivo_existente = novo_arquivo
            print(f"Arquivo criado: {novo_arquivo}")

            # Preencher a célula B3 na aba "Amort. Op." com o código de operação
            wb = load_workbook(arquivo_existente)
            if 'Amort. Op.' in wb.sheetnames:
                sheet_amort_op = wb['Amort. Op.']
                sheet_amort_op['B3'] = ticker
            wb.save(arquivo_existente)
        else:
            # Usar o arquivo existente
            arquivo_existente = os.path.join(diretorio_salvar, arquivo)
            print(f"Arquivo existente usado: {arquivo_existente}")

        # Carregar o workbook
        wb = load_workbook(arquivo_existente)
        if 'Ativos Filtrados' not in wb.sheetnames:
            # Criar a aba "Ativos Filtrados" se não existir
            wb.create_sheet('Ativos Filtrados')
        sheet_ativos_filtrados = wb['Ativos Filtrados']

        # Atualizar a aba "Ativos Filtrados"
        # Remover o cabeçalho ao adicionar as linhas
        for r_idx, r in enumerate(dataframe_to_rows(df_filtrado_operacoes, index=False, header=False), 2):
            for c_idx, value in enumerate(r, 1):
                cell = sheet_ativos_filtrados.cell(row=r_idx, column=c_idx, value=value)
                # Verificar se a coluna é uma das que precisam de formatação de data
                col_name = df_filtrado_operacoes.columns[c_idx - 1]
                if col_name in ["Início Período de Graça", "Fim Período de Graça"]:
                    format_date(cell)

        # Nomear as colunas "Pagamento", "Data Pgmto.", "Data Lançamento Op. 1", "Data Lançamento Op. 2" e "Valor Justo"
        if sheet_ativos_filtrados.cell(row=1, column=30).value is None:
            sheet_ativos_filtrados.cell(row=1, column=30, value='Pagamento')
        if sheet_ativos_filtrados.cell(row=1, column=31).value is None:
            sheet_ativos_filtrados.cell(row=1, column=31, value='Data Pgmto.')
        if sheet_ativos_filtrados.cell(row=1, column=32).value is None:
            sheet_ativos_filtrados.cell(row=1, column=32, value='Data Lançamento Op. 1')
        if sheet_ativos_filtrados.cell(row=1, column=33).value is None:
            sheet_ativos_filtrados.cell(row=1, column=33, value='Data Lançamento Op. 2')
        if sheet_ativos_filtrados.cell(row=1, column=36).value is None:
            sheet_ativos_filtrados.cell(row=1, column=36, value='Valor Justo')

        # Atualizar os valores de pagamento, data, Data Lançamento Op. 1, Data Lançamento Op. 2 e valor justo
        for row_idx, row in enumerate(sheet_ativos_filtrados.iter_rows(min_row=2, max_row=sheet_ativos_filtrados.max_row, min_col=1, max_col=sheet_ativos_filtrados.max_column), 2):
            if row[0].value == nome_cedente:
                sheet_ativos_filtrados.cell(row=row_idx, column=30, value=valor_pago)
                sheet_ativos_filtrados.cell(row=row_idx, column=31, value=data_pagamento)
                sheet_ativos_filtrados.cell(row=row_idx, column=36, value=valor_justo)
            if row_idx - 2 < len(valores_historico):
                cell_data_op1 = sheet_ativos_filtrados.cell(row=row_idx, column=32, value=valores_historico[0] if len(valores_historico) > 0 else None)
                format_date(cell_data_op1)
                if len(valores_historico) > 1:
                    cell_data_op2 = sheet_ativos_filtrados.cell(row=row_idx, column=33, value=valores_historico[1] if len(valores_historico) > 1 else None)
                    format_date(cell_data_op2)

        # Garantir a formatação correta das colunas de data na aba "Ativos Filtrados"
        for row in sheet_ativos_filtrados.iter_rows(min_row=2, max_row=sheet_ativos_filtrados.max_row, min_col=1, max_col=sheet_ativos_filtrados.max_column):
            for cell in row:
                col_name = df_base.columns[cell.column - 1] if cell.column <= len(df_base.columns) else None
                if col_name in ["Início Período de Graça", "Fim Período de Graça", "Data Lançamento Op. 1", "Data Lançamento Op. 2"]:
                    format_date(cell)

        # Criar cópias das abas padrão para cada cedente na aba "Ativos Filtrados"
        abas_padrao = ["Alocação - Ativo", "Flx. Ativo - Real", "Flx. Ativo - Venda"]
        cedentes = []
        for row in sheet_ativos_filtrados.iter_rows(min_row=2, max_row=sheet_ativos_filtrados.max_row, min_col=1, max_col=1):
            if row[0].value:  # Verificar se a célula não é None
                cedentes.append(row[0].value)  # Usar o nome completo do cedente

        # Renomear as abas padrão com o nome do primeiro cedente
        existing_sheet_names = wb.sheetnames
        for cedente in cedentes:
            for aba in abas_padrao:
                primeiro_nome = cedente.split()[0]
                novo_nome = generate_unique_sheet_name(aba.replace('Ativo', primeiro_nome), existing_sheet_names)
                if aba in wb.sheetnames:
                    new_sheet = wb.copy_worksheet(wb[aba])
                    new_sheet.title = novo_nome
                    existing_sheet_names.append(novo_nome)
                    if "Alocação" in novo_nome:
                        new_sheet['B3'] = cedente  # Preencher a célula B3 com o nome completo do cedente
                    if "Flx." in novo_nome:
                        new_sheet['D4'] = cedente  # Preencher a célula D4 com o nome completo do cedente

        wb.save(arquivo_existente)
        print(f'Arquivo "{arquivo_existente}" atualizado com sucesso.')

# Inserir o nome do cedente, valor pago, data do pagamento e valor justo
nome_cedente = input("Digite o nome do cedente: ")
valor_pago = format_currency(input("Digite o valor pago: "))
data_pagamento = input("Digite a data do pagamento: ")
valor_justo = format_currency(input("Digite o valor justo: "))

# Caminho do arquivo Excel de base de dados
caminho_base_dados = r"C:\Users\Gabriel Zambello\Hurst Capital\Legal Claims - Documentos\Asset\EMPIRICUS\Base de Dados.xlsx"

# Caminho da planilha padrão
caminho_planilha_padrao = r"C:\Users\Gabriel Zambello\Hurst Capital\Legal Claims - Documentos\Gestão\Ativos\Controle\Distribuição\Amortização Operações\Planilha Padrao_amort.EMP.xlsx"

try:
    # Ler a aba "Base de Dados Ativos" da planilha Excel
    df_base = pd.read_excel(caminho_base_dados, sheet_name='Base de Dados Ativos')
    filtrar_cedente_e_atualizar_planilha(nome_cedente, valor_pago, data_pagamento, valor_justo, df_base, caminho_planilha_padrao, caminho_base_dados)
except PermissionError:
    print(f"Permissão negada para acessar o arquivo: {caminho_base_dados}")
except FileNotFoundError:
    print(f"Arquivo não encontrado: {caminho_base_dados}")
except KeyError as e:
    print(f"A coluna especificada não foi encontrada: {e}")
except Exception as e:
    print(nome_cedente, valor_pago, data_pagamento, valor_justo)
    print(f"Ocorreu um erro: {e}")
    print(traceback.format_exc())